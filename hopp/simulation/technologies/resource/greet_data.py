import os
import csv
from pathlib import Path
from typing import Union
import requests
import json
import yaml
import openpyxl
import warnings

from hopp.utilities.log import hybrid_logger as logger
# from hopp.simulation.technologies.resource.resource import Resource
from hopp import ROOT_DIR

# Ignore / stop print IO of UserWarnings when opening the GREET excel docs with openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class GREETData:
    """
    Class to manage GREET data

    Args:
        greet_year: (int) version / vintage of GREET to pull data from, Default = 2023
        path_resource: directory where to save data files, Default = hopp_install_dir/hopp/simulation/technologies/resource_files
        filepath: absolute file path of the greet_<year>_processed.yaml to load data from, allows users to manually specify which yaml file to use for adhoc add/edit of values. Default = ""
        preprocess_greet: Flag to preprocess and parse all greet files even if greet_<year>_processed.yaml already exists. Default = False
        kwargs: additional keyword arguments

    """

    def __init__(
        self,
        greet_year: int = 2023,
        path_resource: Union[str, Path] = ROOT_DIR / "simulation" / "resource_files",
        filepath: Union[str, Path] ="",
        preprocess_greet: bool = False,
        **kwargs
    ):

        self.year = greet_year

        # Check if path_resource is a directory, if yes define as self.path_resource attribute
        if os.path.isdir(path_resource):
            self.path_resource = path_resource
        
        # update path with GREET directory
        self.path_resource = os.path.join(self.path_resource, 'greet', str(self.year))

        # Force override internal definitions if kwargs passed in
        self.__dict__.update(kwargs)

        # Define the filepath and filename for the resource file
        if filepath == "":
            filepath = os.path.join(self.path_resource, "greet" + "_" + str(self.year) + "_" + "processed.yaml")
        self.filename = filepath

        # Check if the download directory exists (HOPP/hopp/simulation/resource_files/greet/<year>), if not make the directory
        # NOTE: this directory should always exist as it contains versions of GREET excel hosted in the repo to pull data from
        self.check_download_dir()

        # If a processed resource file does not already exist or preprocess_greet flag == True, process / parse the data from GREET excel docs
        if not os.path.isfile(self.filename) or preprocess_greet:
            self.preprocess_greet()

        # Check if greet_X_processed.yaml exists, if yes load yaml to dictionary and save to self.data, if not error.
        self.format_data()

    def check_download_dir(self):
        if not os.path.isdir(os.path.dirname(self.filename)):
            os.makedirs(os.path.dirname(self.filename))

    def preprocess_greet(self):
        print("************************************************************")
        print("Processing GREET data, this may take up to 2+ minutes")
        ## Define Conversions for GREET data points
        # Unit conversions
        btu_to_kWh = 0.0002931                              # 1 btu = 0.00029307107 kWh
        btu_to_MJ = 0.0010551                               # 1 btu = 0.00105505585262 MJ 
        mmbtu_to_kWh = 293.1                                # 1 MMbtu = 293.07107 kWh 
        mmbtu_to_MWh = 0.2931                               # 1 MMbtu = 0.29307107 MWh
        mmbtu_to_MJ = 1055.1                                # 1 MMbtu = 1055.05585262 MJ
        mmbtu_to_GJ = 1.0551                                # 1 MMbtu = 1.05505585262 GJ
        MJ_to_kWh = 0.2778                                  # 1 MJ = (1/3.6) kWh ~= 0.2777777777777778 kWh
        ton_to_kg = 907.18                                  # 1 US ton = 907.18474 kg
        ton_to_MT = 0.90718                                 # 1 US ton = 0.90718474 metric tonne
        g_to_kg  = 0.001                                    # 1 g = 0.001 kg
        kg_to_MT = 0.001                                    # 1 kg = 0.001 metric tonne

        # Chemical properties
        # mmbtuhhv_per_kg_h2 = 0.134                          # 1 kg H2 = 0.134 MMbtu-hhv h2
        mmbtulhv_per_kg_h2 = 0.114                          # 1 kg H2 = 0.114 MMbtu-lhv H2
        # MJHHV_per_kg_h2 = 141.88                            # Higher Heating Value of hydrogen = 141.88 MJ-HHV/kg H2
        MJLHV_per_kg_h2 = 119.96                            # Lower Heating Value of hydrogen = 119.96 MJ-LHV/kg H2
        # kWh_per_kg_h2_LHV = (MJLHV_per_kg_h2 * MJ_to_kWh)   # kWh per kg of hydrogen using LHV, ~= 33.3222222 kWh/kg H2
        # kWh_per_kg_h2_HHV = (MJHHV_per_kg_h2 * MJ_to_kWh)   # kWh per kg of hydrogen using HHV, ~= 39.4111111 kWh/kg H2
        gal_H2O_to_MT = 0.00378541                          # 1 US gallon of H2O = 0.00378541 metric tonnes (1 gal = 3.78541 liters, 1 liter H2O = 1 kg, 1000 kg = 1 metric tonne)

        # Chemical conversion formulas for greenhouse gases (GHGs) to CO2e emissions intensities (EI) with GWP and Carbon ratios
        # CO2 (VOC, CO, CO2) = CO2 + (VOC*0.85/0.27) + (CO*0.43/0.27)
        # GHGs = CO2 (VOC, CO, CO2) + (CH4*29.8) + (N2O)*273 + (VOC*0) + (CO*0) + (NOx*0) + (BC*0) + (OC*0)
        # GHGs = CO2 + (VOC*0.85/0.27) + (CO*0.43/0.27) + (CH4*29.8) + (N20*273)
        # Carbon Ratios of Pollutants (pulled from GREET1 > Fuel_Specs > B161:B164)
        VOC_to_CO2e = (0.85/0.272727)
        CO_to_CO2e = (0.428571/0.272727)
        CH4_to_CO2e = (0.75/0.272727)
        # Global Warming Potential (relative to CO2, pulled from GREET 1 > Fuel_Specs > B148:B149)
        CH4_gwp_to_CO2e = 29.8          # 1g CH4 == 29.8 g CO2e (not combusted)
        N2O_gwp_to_CO2e = 273           # 1g N2O == 273 g CO2e (not combusted)

        ## Define hardcoded values for efficiencies, emissions intensities (EI), combustion, consumption, and production processes
        # NOTE: In future, update to pull hardcoded values from GREET or other models programmatically if possible
        # Following values determined through communications with GREET / ANL team
        # Water
        H2O_supply_EI = 0.00013         # Emissions Intensity of supplying water to processes (kg CO2e/gal H2O)
        # SMR and ATR
        smr_HEX_eff = 0.9               # Heat exchange efficiency (%)
        ccs_PO_consume = 0              # Power consumption for CCS (kWh/kg CO2)
        atr_steam_prod = 0              # No steam exported during ATR (MJ/kg H2)
        atr_ccs_steam_prod = 0          # No steam exported during ATR (MJ/kg H2)
        # Electrolysis
        # NOTE: following X_ely_PO_consume values not used in current LCA calculations, if used in future these values should come from HOPP/GreenHEART actual electricity consumption of electrolyzer modeled
        pem_ely_PO_consume = 51                                                                     # PEM water electrolysis energy consumption (kWh/kg H2). HFTO 2022 status = 55, 2026 target = 51, Ultimate target = 46 https://www.energy.gov/eere/fuelcells/technical-targets-proton-exchange-membrane-electrolysis
        soec_ely_PO_consume = 44                                                                    # High Temp / Solid Oxide Electrolysis energy consumption (kWh/kg H2). HFTO 2022 status = 47, 2026 target = 44, Ultimate target = 42 https://www.energy.gov/eere/fuelcells/technical-targets-high-temperature-electrolysis
        alk_ely_PO_consume = 52                                                                     # Alkaline water electrolysis energy consumption (kWh/kg H2). HFTO 2022 status = 55, 2026 target = 52, Ultimate target = 48 https://www.energy.gov/eere/fuelcells/technical-targets-liquid-alkaline-electrolysis
        # Battery
        battery_LFP_EI = 20                                                                         # LFP Battery embodied emissions (g CO2e/kWh)

        ## Pull GREET Values
        # NOTE: following logic / cells to pull data from was created for GREET 2023, in future versions may need to add if statements for year / update cells to pull from
        # Define GREET filepaths
        greet1_blue_NH3 = os.path.join(self.path_resource, "blue_NH3_prod", "GREET1_2023_Rev1.xlsm")
        greet2_blue_NH3 = os.path.join(self.path_resource, "blue_NH3_prod", "GREET2_2023_Rev1.xlsm")
        greet1_green_NH3 = os.path.join(self.path_resource, "green_NH3_prod", "GREET1_2023_Rev1.xlsm")
        greet2_green_NH3 = os.path.join(self.path_resource, "green_NH3_prod", "GREET2_2023_Rev1.xlsm")
        greet1_conventional_NH3 = os.path.join(self.path_resource, "conventional_NH3_prod", "GREET1_2023_Rev1.xlsm")
        greet2_conventional_NH3 = os.path.join(self.path_resource, "conventional_NH3_prod", "GREET2_2023_Rev1.xlsm")
        greet1_ccs_central_h2 = os.path.join(self.path_resource, "ccs_central_h2_prod", "GREET1_2023_Rev1.xlsm")
        greet2_ccs_central_h2 = os.path.join(self.path_resource, "ccs_central_h2_prod", "GREET2_2023_Rev1.xlsm")
        greet1_no_ccs_central_h2 = os.path.join(self.path_resource,"no_ccs_central_h2_prod", "GREET1_2023_Rev1.xlsm")
        greet2_no_ccs_central_h2 = os.path.join(self.path_resource,"no_ccs_central_h2_prod", "GREET2_2023_Rev1.xlsm")

        #------------------------------------------------------------------------------
        # Renewable infrastructure embedded emission intensities, Lime and Natural Gas emission intensities, Ammonia, Hydrogen production via water electrolysis, and Steel
        #------------------------------------------------------------------------------
        # NOTE: Capex EI, Lime and NG EI, Ammonia chemical consumption, Electrolysis, and Steel GREET values agnostic of ccs/no_ccs and NH3 production methods, ie: can use any version of greet to pull
        # NOTE: For Steel, alternative DRI-EAF configurations (w/ and w/out scrap, H2 vs NG) found in greet2 > Steel > W107:Z136
                # Iron ore vs scrap % controlled by B24:C24 values
                    # May require hosting of different steel config GREET versions if desired
                # Values below are for DRI-EAF 83% H2, 100% DRI 0% Scrap
        greet1 = openpyxl.load_workbook(greet1_ccs_central_h2, data_only=True)
        # Renewable Infrastructure
        wind_capex_EI = (greet1['ElecInfra']['G112'].value / mmbtu_to_kWh)                          # Wind CAPEX emissions (g CO2e/kWh)
        solar_pv_capex_EI = (greet1['ElecInfra']['H112'].value / mmbtu_to_kWh)                      # Solar PV CAPEX emissions (g CO2e/kWh)
        nuclear_PWR_capex_EI = (greet1['ElecInfra']['D112'].value / mmbtu_to_kWh)                   # Nuclear Pressurized Water Reactor (PWR) CAPEX emissions (g CO2e/kWh)
        nuclear_BWR_capex_EI = (greet1['ElecInfra']['E112'].value / mmbtu_to_kWh)                   # Nuclear Boiling Water Reactor (BWR) CAPEX emissions (g CO2e/kWh)
        coal_capex_EI = (greet1['ElecInfra']['B112'].value / mmbtu_to_kWh)                          # Coal CAPEX emissions (g CO2e/kWh)
        gas_capex_EI = (greet1['ElecInfra']['C112'].value / mmbtu_to_kWh)                           # Natural Gas Combined Cycle (NGCC) CAPEX emissions (g CO2e/kWh)
        hydro_capex_EI = (greet1['ElecInfra']['F112'].value / mmbtu_to_kWh)                         # Hydro CAPEX emissions (g CO2e/kWh)
        bio_capex_EI = (greet1['ElecInfra']['L112'].value / mmbtu_to_kWh)                           # Biomass CAPEX emissions (g CO2e/kWh)
        geothermal_EGS_capex_EI = (greet1['ElecInfra']['I112'].value / mmbtu_to_kWh)                # Geothermal EGS CAPEX emissions (g CO2e/kWh)
        geothermal_flash_capex_EI = (greet1['ElecInfra']['J112'].value / mmbtu_to_kWh)              # Geothermal Flash CAPEX emissions (g CO2e/kWh)
        geothermal_binary_capex_EI = (greet1['ElecInfra']['K112'].value / mmbtu_to_kWh)             # Geothermal Binary CAPEX emissions (g CO2e/kWh)
        # Lime
        lime_supply_EI = ((greet1['Chemicals']['BA247'].value +                                     # GHG Emissions Intensity of supplying Lime to processes accounting for limestone mining, lime production, lime processing, and lime transportation assuming 20 miles transport via Diesel engines (kg CO2e/kg lime)
                         (greet1['Chemicals']['BA237'].value * VOC_to_CO2e) +             
                         (greet1['Chemicals']['BA238'].value * CO_to_CO2e) +
                         (greet1['Chemicals']['BA245'].value * CH4_gwp_to_CO2e) +
                         (greet1['Chemicals']['BA246'].value * N2O_gwp_to_CO2e)
                         ) * g_to_kg * (1/ton_to_kg))
        # Natural Gas (NG)
        NG_combust_EI = ((greet1['EF']['B16'].value +                                               # GHG Emissions Intensity of Natural Gas combustion in a utility / industrial large boiler (g CO2e/MJ Natural Gas combusted)
                         (greet1['EF']['B6'].value * VOC_to_CO2e) +             
                         (greet1['EF']['B7'].value * CO_to_CO2e) +
                         (greet1['EF']['B14'].value * CH4_gwp_to_CO2e) +
                         (greet1['EF']['B15'].value * N2O_gwp_to_CO2e)
                         )/mmbtu_to_MJ)
        
        NG_supply_EI = (greet1['NG']['B103'].value / mmbtu_to_MJ)                                   # GHG Emissions Intensity of supplying Natural Gas to processes as a feedstock or process fuel (g CO2e/MJ Natural Gas consumed)
                                                                                                    # NOTE: this values assumes CH4 leakage rates as described in GREET1 > Inputs > P127:Q135
                                                                                                    # NOTE: this value assumes a mix of 25% conventional NG and 75% shale gas for North American NG supply (controlled by GREET1 > Inputs > F120:F121)
                                                                                                    # NOTE: this value can account for importation of NG from overseas depending on value set in GREET1 > NG > B118, currently set to 0% import from overseas
        # Ammonia (NH3)
        NH3_NG_consume = ((greet1['Ag_Inputs']['F41'].value *                                       # Natural gas consumption for combustion in the Haber-Bosch process / Boiler for Ammonia production (MJ/metric tonne NH3) 
                           greet1['Ag_Inputs']['F44'].value) *
                           mmbtu_to_MJ/ton_to_MT)       
        NH3_H2_consume =  (greet1['Ag_Inputs']['AM54'].value)                                       # Gaseous Hydrogen consumption for Ammonia production, based on chemical balance and is applicable for all NH3 production pathways (kg H2/kg NH3)
        NH3_electricity_consume = (greet1['Ag_Inputs']['F45'].value *                               # Total Electrical Energy consumption for Ammonia production, based on chemical / energy balance and is applicable for all NH3 productin pathways (kWh/kg NH3)
                                   mmbtu_to_kWh/ton_to_kg)                                          # NOTE: additional energy consumption for h2 production via electrolysis (green NH3) and carbon capture (blue NH3) are accounted for in other variables

        greet1.close()

        greet2 = openpyxl.load_workbook(greet2_ccs_central_h2, data_only=True)
        # Renewable Infrastructure
        # battery_LFP_residential_EI = (greet2['Solar_PV']['DI289'].value)                            # NOTE: original value = 20 (no indication of residential or commercial), greet value = 172.92 (30 yrs, 100 MWh battery), Battery embodied emissions for residential solar PV applications (g CO2e/kWh), assumed LFP batteries (can update battery chemistry in cell K155)
        #                                                                                                 # value must be multiplied by factor = (project_lifetime / battery_system_capacity_kwh) before use in LCA calculations, opted not to divide here so value remains agnostic of project lifetime and system battery size        
        # battery_LFP_commercial_EI = (greet2['Solar_PV']['DJ289'].value)                             # NOTE: original value = 20 (no indication of residential or commercial), greet value = 106362.41 (30 yrs, 100 MWh battery), Battery embodied emissions for commercial solar PV applications (g CO2e/kWh), assumed LFP batteries (can update battery chemistry in cell W155)
        #                                                                                                 # value must be multiplied by factor = (project_lifetime / battery_system_capacity_kwh) before use in LCA calculations, opted not to divide here so value remains agnostic of project lifetime and system battery size
        # Electrolysis 
        pem_ely_stack_capex_EI = (greet2['Electrolyzers']['I257'].value * g_to_kg)                  # PEM electrolyzer stack CAPEX emissions (kg CO2e/kg H2)
        pem_ely_stack_and_BoP_capex_EI = (greet2['Electrolyzers']['L257'].value * g_to_kg)          # PEM electrolyzer stack + Balance of Plant CAPEX emissions (kg CO2e/kg H2)
        alk_ely_stack_capex_EI = (greet2['Electrolyzers']['O257'].value * g_to_kg)                  # Alkaline electrolyzer stack CAPEX emissions (kg CO2e/kg H2)
        alk_ely_stack_and_BoP_capex_EI = (greet2['Electrolyzers']['R257'].value * g_to_kg)          # Alkaline electrolyzer stack + Balance of Plant CAPEX  emissions (kg CO2e/kg H2)
        soec_ely_stack_capex_EI = (greet2['Electrolyzers']['C257'].value * g_to_kg)                 # SOEC electrolyzer stack CAPEX emissions (kg CO2e/kg H2)
        soec_ely_stack_and_BoP_capex_EI = (greet2['Electrolyzers']['F257'].value * g_to_kg)         # SOEC electrolyzer stack + Balance of Plant CAPEX emissions (kg CO2e/kg H2)
        # Steel
        steel_CH4_prod = (greet2['Steel']['Y123'].value * CH4_gwp_to_CO2e * g_to_kg / ton_to_MT)                    # TODO: See GreenHEART_Notes.txt, CH4 emissions for DRI-EAF Steel production w/ 83% H2 and 0% scrap (kg CO2e/metric tonne annual steel lab production)
        steel_CO2_prod = (greet2['Steel']['Y125'].value * g_to_kg / ton_to_MT)                                      # TODO: See GreenHEART_Notes.txt, CO2 emissions for DRI-EAF Steel production w/ 83% H2 and 0% scrap (kg CO2e/metric tonne annual steel lab production)
        steel_iron_ore_EI = ((greet2['Steel']['B92'].value +                                                        # Iron ore production GHG emissions for use in DRI-EAF Steel production (kg CO2e/kg iron ore)
                             (greet2['Steel']['B82'].value * VOC_to_CO2e) +                     
                             (greet2['Steel']['B83'].value * CO_to_CO2e) +
                             (greet2['Steel']['B90'].value * CH4_gwp_to_CO2e) + 
                             (greet2['Steel']['B91'].value * N2O_gwp_to_CO2e)
                             ) * (g_to_kg / ton_to_kg))
        steel_H2O_consume = (greet2['Steel']['Y113'].value * (gal_H2O_to_MT/ton_to_MT))                             # TODO: See GreenHEART_Notes.txt, H2O consumption for DRI-EAF Steel production w/ 83% H2 and 0% scrap (metric tonne H2O/metric tonne steel production)
        steel_H2_consume = (greet2['Steel']['AK66'].value * (mmbtu_to_MJ/MJLHV_per_kg_h2) * (kg_to_MT/ton_to_MT))   # Hydrogen consumption for DRI-EAF Steel production w/ 83% H2 regardless of scrap (metric tonnes H2/metric tonne steel production)
        steel_NG_consume = ((greet2['Steel']['AE63'].value +                                                        # Natural gas consumption for DRI-EAF Steel production accounting for iron ore mining, pelletization, DRI with 83% H2, and EAF + LRF (GJ/metric tonne steel)
                             greet2['Steel']['AG63'].value +                                    
                             greet2['Steel']['AK63'].value + 
                             greet2['Steel']['AM63'].value
                            ) * (mmbtu_to_GJ / ton_to_MT))
        steel_lime_consume = (greet2['Steel']['AM68'].value)                                                        # Lime consumption for DRI-EAF Steel production (metric tonne lime/metric tonne steel production)
        steel_iron_ore_consume = (greet2['Steel']['AM69'].value)                                                    # Iron ore consumption for DRI-EAF Steel production (metric tonne iron ore/metric tonne steel production)
        steel_electricity_consume_scope3 = ((greet2['Steel']['AE65'].value +                                        # Scope 3 related Total Electrical Energy consumption for DRI-EAF Steel production accounting for iron ore mining, pelletization, DRI with 83% H2, and EAF + LRF (MWh/metric tonne steel production)
                                             greet2['Steel']['AG65'].value +
                                             greet2['Steel']['AK65'].value + 
                                             greet2['Steel']['AM65'].value
                                             ) * (mmbtu_to_MWh/ton_to_MT))
        steel_electricity_consume_scope2 = ((greet2['Steel']['AK65'].value +                                        # Scope 2 related Total Electrical Energy consumption for DRI-EAF Steel production accounting for DRI with 83% H2 and EAF + LRF (MWh/metric tonne steel production)
                                             greet2['Steel']['AM65'].value
                                             ) * (mmbtu_to_MWh/ton_to_MT))
        greet2.close()

        #------------------------------------------------------------------------------
        # Steam methane reforming (SMR) and Autothermal Reforming (ATR) - Incumbent H2 production processes
        #------------------------------------------------------------------------------
        # Values with Carbon Capture Sequestration (CCS)
        greet1 = openpyxl.load_workbook(greet1_ccs_central_h2, data_only=True)
        smr_ccs_total_energy = (1E6 * greet1['Hydrogen']['B214'].value / greet1['Hydrogen']['B212'].value)       # Total Energy required for H2 production from SMR w/ CCS accounting for loss factor and efficiency (btu/mmbtu H2)
        smr_ccs_steam_prod = (greet1['Inputs']['I1105'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                  # SMR Steam exported w/ CCS (MJ/kg H2)
        smr_ccs_perc_capture = (greet1['Hydrogen']['B11'].value)                                                 # CCS rate for SMR (%)
        atr_ccs_perc_capture = (greet1['Hydrogen']['B15'].value)                                                 # CCS rate for Autothermal Reforming (%)
        # SMR with conventional Natural Gas
        smr_ccs_NG_consume = ((smr_ccs_total_energy *                                                            # Natural gas consumption for SMR w/ CCS accounting for loss factor, efficiency, NG as feed and NG as process fuel (MJ-LHV/kg H2)
                             (greet1['Hydrogen']['B215'].value + 
                             ((1 - greet1['Hydrogen']['B215'].value) * greet1['Hydrogen']['B221'].value))) *
                             btu_to_MJ*mmbtulhv_per_kg_h2)
        smr_ccs_NG_PO_consume = (greet1['Hydrogen']['C389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)                # TODO: See GreenHEART_Notes.txt, SMR via NG w/ CCS Total Energy consumption (kWh/kg H2) 
        # SMR with Renewable Natural Gas
        # smr_ccs_RNG_consume = (greet1['Hydrogen']['L392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                   # TODO: See GreenHEART_Notes.txt, SMR w/ CCS WTG Renewable Natural Gas (RNG) consumption (MJ-LHV/kg H2)
        # smr_ccs_RNG_PO_consume = (greet1['Hydrogen']['L389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)               # TODO: See GreenHEART_Notes.txt, SMR via RNG w/ CCS Total Energy consumption (kWh/kg H2)
        # ATR with conventional Natural Gas
        atr_ccs_NG_consume = (greet1['Hydrogen']['N392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                    # TODO: See GreenHEART_Notes.txt, ATR w/ CCS WTG NG consumption (MJ-LHV/kg H2)
        atr_ccs_NG_PO_consume = (greet1['Hydrogen']['N389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)                # TODO: See GreenHEART_Notes.txt, ATR via NG w/ CCS WTG Total Energy consumption (kWh/kg H2)
        # ATR with Renewable Natural Gas
        # atr_ccs_RNG_consume = (greet1['Hydrogen']['O392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                   # TODO: See GreenHEART_Notes.txt, ATR w/ CCS WTG RNG consumption (MJ-LHV/kg H2)
        # atr_ccs_RNG_PO_consume = (greet1['Hydrogen']['O389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)               # TODO: See GreenHEART_Notes.txt, ATR via RNG w/ CCS WTG Total Energy consumption (kWh/kg H2)
        greet1.close()

        # Values without CCS
        greet1 = openpyxl.load_workbook(greet1_no_ccs_central_h2, data_only=True)
        smr_total_energy = (1E6 * greet1['Hydrogen']['B214'].value / greet1['Hydrogen']['B212'].value)      # Total Energy required for H2 production via SMR w/out CCS accounting for loss factor and efficiency (btu/mmbtu H2)
        smr_steam_prod = (greet1['Inputs']['I1105'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                 # SMR Steam exported w/out CCS (MJ/kg H2)
        # SMR with conventional NG
        smr_NG_consume = ((smr_total_energy *                                                               # Natural gas consumption for SMR w/out CCS accounting for loss factor, efficiency, NG as feed and NG as process fuel (MJ-LHV/kg H2)
                         (greet1['Hydrogen']['B215'].value + 
                         ((1 - greet1['Hydrogen']['B215'].value) * greet1['Hydrogen']['B221'].value))) *
                         btu_to_MJ*mmbtulhv_per_kg_h2)
        smr_NG_PO_consume = (greet1['Hydrogen']['C389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)            # TODO: See GreenHEART_Notes.txt, greet value = 10.02, SMR via NG w/out CCS WTG Total Energy consumption (kWh/kg H2)
        # SMR with Renewable Natural Gas
        # smr_RNG_consume = (greet1['Hydrogen']['L392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)               # TODO: See GreenHEART_Notes.txt, SMR w/out CCS WTG RNG consumption (MJ-LHV/kg H2)
        # smr_RNG_PO_consume = (greet1['Hydrogen']['L389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)           # TODO: See GreenHEART_Notes.txt, SMR via RNG w/out CCS WTG Total Energy consumption (kWh/kg H2)
        #TODO: remove ATR values below, no pathway for ATR without CCS
        # atr_NG_consume = (greet1['Hydrogen']['N392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)                # TODO: See GreenHEART_Notes.txt, same value as w/ CCS enabled, no pathway for ATR w/out CCS, ATR w/out CCS WTG NG consumption (MJ-LHV/kg H2)
        # atr_RNG_consume = (greet1['Hydrogen']['O392'].value * btu_to_MJ * mmbtulhv_per_kg_h2)               # TODO: See GreenHEART_Notes.txt, same value as w/ CCS enabled, no pathway for ATR w/out CCS, ATR w/out CCS WTG RNG consumption (MJ-LHV/kg H2)
        # atr_NG_PO_consume = (greet1['Hydrogen']['N389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)            # TODO: See GreenHEART_Notes.txt, same value as w/ CCS enabled, no pathway for ATR w/out CCS, ATR via NG w/out CCS WTG Total Energy consumption (kWh/kg H2)
        # atr_RNG_PO_consume = (greet1['Hydrogen']['O389'].value * btu_to_kWh * mmbtulhv_per_kg_h2)           # TODO: See GreenHEART_Notes.txt, same value as w/ CCS enabled, no pathway for ATR w/out CCS # ATR via RNG w/out CCS WTG Total Energy consumption (kWh/kg H2)
        greet1.close()

        #------------------------------------------------------------------------------
        # Ammonia (NH3)
        #------------------------------------------------------------------------------
        # # Values for Green NH3
        # greet1 = openpyxl.load_workbook(greet1_green_NH3, data_only=True)
        # green_NH3_PO_consume = (greet1['Hydrogen']['B359'].value * btu_to_kWh)                      # TODO: See GreenHEART_Notes.txt, Green Ammonia production Total Energy consumption (kWh/kg NH3)
        # greet1.close()

        # # Values for Blue NH3
        # greet1 = openpyxl.load_workbook(greet1_blue_NH3, data_only=True)
        # blue_NH3_PO_consume = (greet1['Hydrogen']['B359'].value * btu_to_kWh)                       # TODO: See GreenHEART_Notes.txt, Blue Ammonia production Total Energy consumption (kWh/kg NH3)
        # greet1.close()

        # # Values for Conventional NH3
        # greet1 = openpyxl.load_workbook(greet1_conventional_NH3, data_only=True)
        # conventional_NH3_PO_consume = (greet1['Hydrogen']['B359'].value * btu_to_kWh)               # TODO: See GreenHEART_Notes.txt, Conventional Ammonia production Total Energy consumption (kWh/kg NH3)
        # greet1.close()

        data_dict = {
                     # Hardcoded values
                     'smr_HEX_eff':smr_HEX_eff,
                     'ccs_PO_consume':ccs_PO_consume,
                     'atr_steam_prod':atr_steam_prod,
                     'atr_ccs_steam_prod':atr_ccs_steam_prod,
                     # Natural gas
                     'NG_combust_EI':NG_combust_EI,
                     'NG_supply_EI':NG_supply_EI,
                     # Water
                     'H2O_supply_EI':H2O_supply_EI,
                     # Lime
                     'lime_supply_EI':lime_supply_EI,
                     # Renewable infrastructure embedded EI and h2 production via water electrolysis
                     'wind_capex_EI':wind_capex_EI,
                     'solar_pv_capex_EI':solar_pv_capex_EI,
                     'nuclear_PWR_capex_EI':nuclear_PWR_capex_EI,
                     'nuclear_BWR_capex_EI':nuclear_BWR_capex_EI,
                     'coal_capex_EI':coal_capex_EI,
                     'gas_capex_EI':gas_capex_EI,
                     'hydro_capex_EI':hydro_capex_EI,
                     'bio_capex_EI':bio_capex_EI,
                     'geothermal_EGS_capex_EI':geothermal_EGS_capex_EI,
                     'geothermal_flash_capex_EI':geothermal_flash_capex_EI,
                     'geothermal_binary_capex_EI':geothermal_binary_capex_EI,
                     'pem_ely_PO_consume':pem_ely_PO_consume,
                     'pem_ely_stack_capex_EI':pem_ely_stack_capex_EI,
                     'pem_ely_stack_and_BoP_capex_EI':pem_ely_stack_and_BoP_capex_EI,
                     'alk_ely_PO_consume':alk_ely_PO_consume,
                     'alk_ely_stack_capex_EI':alk_ely_stack_capex_EI,
                     'alk_ely_stack_and_BoP_capex_EI':alk_ely_stack_and_BoP_capex_EI,
                     'soec_ely_PO_consume':soec_ely_PO_consume,
                     'soec_ely_stack_capex_EI':soec_ely_stack_capex_EI,
                     'soec_ely_stack_and_BoP_capex_EI':soec_ely_stack_and_BoP_capex_EI,
                     'battery_LFP_EI':battery_LFP_EI,
                     # Steam methane reforming (SMR) and Autothermal Reforming (ATR)
                     'smr_ccs_NG_consume':smr_ccs_NG_consume,
                     'smr_ccs_RNG_consume':smr_ccs_RNG_consume,
                     'smr_ccs_NG_PO_consume':smr_ccs_NG_PO_consume,
                     'smr_ccs_RNG_PO_consume':smr_ccs_RNG_PO_consume,
                     'smr_ccs_steam_prod':smr_ccs_steam_prod,
                     'smr_ccs_perc_capture':smr_ccs_perc_capture,
                     'atr_ccs_NG_consume':atr_ccs_NG_consume,
                     'atr_ccs_RNG_consume':atr_ccs_RNG_consume,
                     'atr_ccs_NG_PO_consume':atr_ccs_NG_PO_consume,
                     'atr_ccs_RNG_PO_consume':atr_ccs_RNG_PO_consume,
                     'atr_ccs_perc_capture':atr_ccs_perc_capture,
                     'smr_NG_consume':smr_NG_consume,
                     'smr_RNG_consume':smr_RNG_consume,
                     'smr_NG_PO_consume':smr_NG_PO_consume,
                     'smr_RNG_PO_consume':smr_RNG_PO_consume,
                     'smr_steam_prod':smr_steam_prod,
                     'atr_NG_consume':atr_NG_consume,
                     'atr_RNG_consume':atr_RNG_consume,
                     'atr_NG_PO_consume':atr_NG_PO_consume,
                     'atr_RNG_PO_consume':atr_RNG_PO_consume,
                     # Ammonia (NH3)
                     'NH3_NG_consume':NH3_NG_consume,
                     'NH3_H2_consume':NH3_H2_consume,
                     'NH3_electricity_consume':NH3_electricity_consume,
                     # Steel
                     'steel_CH4_prod':steel_CH4_prod,
                     'steel_CO2_prod':steel_CO2_prod,
                     'steel_iron_ore_EI':steel_iron_ore_EI,
                     'steel_H2O_consume':steel_H2O_consume,
                     'steel_H2_consume':steel_H2_consume,
                     'steel_NG_consume':steel_NG_consume,
                     'steel_lime_consume':steel_lime_consume,
                     'steel_iron_ore_consume':steel_iron_ore_consume,
                     'steel_electricity_consume_scope3':steel_electricity_consume_scope3,
                     'steel_electricity_consume_scope2':steel_electricity_consume_scope2,
                    }
        
        # Dump data to yaml file
        yaml_file = open(self.filename, mode="w+")
        yaml.dump(data_dict, yaml_file, default_flow_style=False)
        yaml_file.close()
        print("GREET processing complete")

    def format_data(self):
        """
        """
        if not os.path.isfile(self.filename):
            raise FileNotFoundError(f"{self.filename} does not exist. Try `preprocess_greet` first or provide the absolute path to the greet_<year>_processed.yaml to load.")

        yaml_file = open(self.filename, mode='r')
        self.data = yaml.load(yaml_file, Loader=yaml.SafeLoader)
        yaml_file.close()

#Adhoc testing
if __name__ == '__main__':
    test = GREETData()
    print(test.data)

#NOTE: Runtime ~ 1m16s to fully parse greet
#NOTE: Runtime <2s with greet preprocessed (load yaml to dict)